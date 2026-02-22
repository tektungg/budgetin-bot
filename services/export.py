"""
Export service — generate file Excel (.xlsx) yang dikirim langsung ke chat Telegram.
Dengan header berwarna, kolom rapi, dan summary di bawah.
"""

import io
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from services.parser import format_rupiah

logger = logging.getLogger(__name__)

MONTH_NAMES = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}

# ── Styles ──────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

MASUK_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
KELUAR_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

SUMMARY_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
SUMMARY_FONT = Font(name="Calibri", bold=True, size=11)

NORMAL_FONT = Font(name="Calibri", size=11)
MONEY_FORMAT = '#,##0'

THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

COLUMN_WIDTHS = {
    "A": 5,   # No
    "B": 18,  # Tanggal
    "C": 14,  # Tipe
    "D": 18,  # Jumlah
    "E": 20,  # Kategori
    "F": 30,  # Deskripsi
    "G": 10,  # Sumber
}


def generate_excel(transactions: list[dict], month_name: str, year: int) -> io.BytesIO:
    """
    Generate file Excel dengan styling dari list transaksi.
    Return BytesIO object yang siap dikirim ke Telegram.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year}"

    # ── Set column widths ───────────────────────────────
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Header row ──────────────────────────────────────
    headers = ["No", "Tanggal", "Tipe", "Jumlah (Rp)", "Kategori", "Deskripsi", "Sumber"]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Data rows ───────────────────────────────────────
    for i, tx in enumerate(transactions, 1):
        # Parse tanggal
        created = tx.get("created_at", "")
        if isinstance(created, str):
            from utils.formatter import parse_iso_date, format_tanggal_waktu
            dt = parse_iso_date(created)
            date_str = format_tanggal_waktu(dt) if dt else str(created)
        elif hasattr(created, "strftime"):
            from utils.formatter import format_tanggal_waktu
            date_str = format_tanggal_waktu(created)
        else:
            date_str = ""

        is_masuk = tx["type"] == "masuk"
        tipe = "Pemasukan" if is_masuk else "Pengeluaran"
        row_fill = MASUK_FILL if is_masuk else KELUAR_FILL

        row_data = [
            i,
            date_str,
            tipe,
            tx["amount"],
            tx.get("category", "Lainnya"),
            tx.get("description", ""),
            tx.get("source", "text"),
        ]
        ws.append(row_data)

        row_num = i + 1  # +1 karena header di row 1
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = row_fill
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

            # Alignment
            if col_num == 1:  # No
                cell.alignment = Alignment(horizontal="center")
            elif col_num == 4:  # Jumlah
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right")
            elif col_num in (2, 3, 7):  # Tanggal, Tipe, Sumber
                cell.alignment = Alignment(horizontal="center")

    # ── Summary rows ────────────────────────────────────
    total_masuk = sum(t["amount"] for t in transactions if t["type"] == "masuk")
    total_keluar = sum(t["amount"] for t in transactions if t["type"] == "keluar")
    saldo = total_masuk - total_keluar

    summary_start = len(transactions) + 3  # +1 header, +1 empty row

    summary_data = [
        ("💰 TOTAL PEMASUKAN", total_masuk),
        ("💸 TOTAL PENGELUARAN", total_keluar),
        ("📊 SALDO", saldo),
    ]

    for idx, (label, value) in enumerate(summary_data):
        row = summary_start + idx
        ws.cell(row=row, column=3, value=label)
        ws.cell(row=row, column=4, value=value)

        for col in range(3, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = SUMMARY_FILL
            cell.font = SUMMARY_FONT
            cell.border = THIN_BORDER
            if col == 4:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right")

    # ── Save to buffer ──────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = f"Budgetin_{month_name}_{year}.xlsx"

    logger.info(f"Generated Excel: {len(transactions)} transactions")
    return buffer


def build_export_caption(transactions: list[dict], month_name: str, year: int) -> str:
    """Buat caption untuk file yang dikirim"""
    total_masuk = sum(t["amount"] for t in transactions if t["type"] == "masuk")
    total_keluar = sum(t["amount"] for t in transactions if t["type"] == "keluar")
    saldo = total_masuk - total_keluar
    saldo_emoji = "📈" if saldo >= 0 else "📉"

    return (
        f"📤 <b>Export {month_name} {year}</b>\n\n"
        f"📝 {len(transactions)} transaksi\n"
        f"💰 Pemasukan: <b>{format_rupiah(total_masuk)}</b>\n"
        f"💸 Pengeluaran: <b>{format_rupiah(total_keluar)}</b>\n"
        f"{saldo_emoji} Saldo: <b>{format_rupiah(abs(saldo))}</b>"
    )
